import pandas as pd
import openpyxl

table = pd.read_html("sale_order.xls")
table[0].to_excel("data_refresh.xlsx", index=False, header=False)
excelFile = openpyxl.load_workbook("data_refresh.xlsx")
sheet1 = excelFile.active
COL_ALPHABET = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L")


# Функция, в которую передаётся название необходимого столбца (col_name). Далее передаётся текст, который необходимо
# заменить в каждой ячейке столбца (what_change_to) на значение текста, которое вам нужно (what_change_for)


def s_and_r_in_col(col_name, what_change_to, what_change_for):
    if sheet1[COL_ALPHABET[col] + "1"].value == col_name:
        for row in range(2, sheet1.max_row + 1):
            sheet1[COL_ALPHABET[col] + str(row)] = sheet1[COL_ALPHABET[col] + str(row)].value.replace(what_change_to,
                                                                                                  what_change_for)


# Функция, в которую передаётся название необходимого столбца (col_name_0), а также символ, по которому будет
# разделяться содержимое ячейки (разделить, split_sybmol). Часть до разделителя пишется в исходную ячейку, часть после
# разделителя пишется в ячейку справа от исходной.


def s_and_split_in_2_cols(col_name_0, split_symbol):
    if sheet1[COL_ALPHABET[col] + "1"].value == col_name_0:
        for row in range(2, sheet1.max_row + 1):
            date_order = sheet1[COL_ALPHABET[col] + str(row)].value.split().pop().split(split_symbol)
            sheet1[COL_ALPHABET[col] + str(row)] = date_order[0]
            sheet1[COL_ALPHABET[col + 1] + str(row)] = date_order[1]


for col in range(0, sheet1.max_column + 1):
    s_and_r_in_col("ID", "№", "")
    s_and_r_in_col("Сумма", ".00 руб.", "")
    s_and_split_in_2_cols("Дата и время доставки", "-")
excelFile.save("data.xlsx")
s_and_r_in_col("ID", "№", "")
