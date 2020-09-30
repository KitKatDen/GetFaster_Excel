import openpyxl

excelFile = openpyxl.load_workbook('report.xlsx')
sheets = excelFile.sheetnames
COL_LETTER = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L')
money_from_clients = False
sum_check_bool = False
canceled_orders = 0
not_taken_IQOS = 0
cash = 0
cc = 0
sum_not_delivered = 0

for i in range(len(sheets)):
    # шаг по листам
    active_sheet = excelFile[sheets[i]]
    nq_orders_check = False     # проверка наличия пункта "Всего заказов"
    for n in range(0, active_sheet.max_column + 1):
        # шаг по столбцам
        for n1 in range(1, active_sheet.max_row + 1):
            #шаг по строчкам
            if str(active_sheet[COL_LETTER[n] + str(n1)].value).find('Всего заказов') == 0:
                n_orders = active_sheet[COL_LETTER[n + 1] + str(n1)].value  # количество заказов у курьера
                nq_orders_check = True  # проверка наличия пункта "Всего заказов"
                cancl_orders_check = False  # проверка наличия пункта "Отменен"
            if str(active_sheet[COL_LETTER[n] + str(n1)].value).find('Отменен') == 0 and cancl_orders_check == False:
                canceled_orders = active_sheet[
                    COL_LETTER[n + 1] + str(n1)].value  # количество отменённых заказов у курьера
                cancl_orders_check = True
            if str(active_sheet[COL_LETTER[n] + str(n1)].value).find('Отменен') == 0 and active_sheet[
                COL_LETTER[n + 1] + str(n1)].value != canceled_orders and active_sheet[
                COL_LETTER[n + 1] + str(n1)].value == '-':
                not_taken_IQOS = not_taken_IQOS + 1
                #if n-1 >= 1 and str(active_sheet[col_letter[n] + str(n1)].value).find('Отменен') == 0:
                #  print(active_sheet[col_letter[n-1] + str(n1)].value)
                #  sum_not_delivered = sum_not_delivered + int(active_sheet[col_letter[n-1] + str(n1)].value)
            if str(active_sheet[COL_LETTER[n] + str(n1)].value).find('Получено денег (суммарная стоимость)') == 0:
                sum_check = int(active_sheet[COL_LETTER[n + 1] + str(n1)].value.split(' из ').pop().split('.')[0].split('(').pop())
                sum_check_bool = True
        if active_sheet[COL_LETTER[n] + '2'].value == 'Получено с клиента':
            money_from_clients = True  # проверка наличия пункта "Получено с клиента"
            cash = 0  # сумма налички
            q_cash = 0  # кол-во заказов оплаченных наличкой
            cc = 0  # сумма карты
            q_cc = 0  # кол-во заказов оплаченных картой
            for p in range(1, n_orders + 1):
                if_cash = active_sheet[COL_LETTER[n + 2] + str(2 + p)].value
                if_cc = active_sheet[COL_LETTER[n + 2] + str(2 + p)].value
                if if_cash.find('Нал') == 0 or if_cash.find('Нал') == 1 or if_cash.find('нал') == 0 or if_cash.find(
                        'нал') == 1:
                    cash = cash + active_sheet[COL_LETTER[n] + str(2 + p)].value
                    q_cash = q_cash + 1
                if if_cc.find('Карта') == 0 or if_cc.find('Карта') == 1 or if_cc.find('карта') == 0 or if_cc.find(
                        'карта') == 1:
                    cc = cc + active_sheet[COL_LETTER[n] + str(2 + p)].value
                    q_cc = q_cc + 1
    total_cash = 0
    total_cc = 0
    total_cash = total_cash + cash
    total_cc = total_cc + cc
#if 1==1: #sum_check == cash + cc+sum_not_delivered:
#print(sum_not_delivered, cash, cc)
    if not nq_orders_check:
        print('Ошибка, не найдено количество заказов')
    if not money_from_clients:
        print('Ошибка, не найдено ячейки "Получено с клиента"')
    if q_cc <= 4 and q_cc != 1 and q_cc != 0:
        print(sheets[i] + ': ' + str(cash) + ' наличными, ' + str(q_cc) + ' чека, ' + str(
            canceled_orders - not_taken_IQOS) + ' IQOS')
    elif q_cc == 1:
        print(sheets[i] + ': ' + str(cash) + ' наличными, ' + str(q_cc) + ' чек, ' + str(
            canceled_orders - not_taken_IQOS) + ' IQOS')
    elif q_cc == 0:
        print(sheets[i] + ': ' + str(cash) + ' наличными, ' + str(q_cc) + ' чеков, ' + str(
            canceled_orders - not_taken_IQOS) + ' IQOS')
    else:
        print(sheets[i] + ': ' + str(cash) + ' наличными, ' + str(q_cc) + ' чеков, ' + str(
            canceled_orders - not_taken_IQOS) + ' IQOS')
#print('Ошибка, не сходятся суммы, ')
print('За сегодня ' + str(total_cash) + ' наличными, ' + str(total_cc) + ' по карте')